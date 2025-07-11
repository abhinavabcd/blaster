'''
Created on 04-Nov-2017

@author: abhinav
'''

from gevent.threading import Thread
from gevent.queue import Queue, Empty as QueueEmptyException
import os
import sys
import csv
import weakref
import subprocess
import shlex
import collections
import random
import string
import socket
import struct
import fcntl
import heapq
import types
from gevent import sleep
from functools import reduce as _reduce
from gevent.lock import BoundedSemaphore
from datetime import timezone, timedelta, datetime
import time
import hmac
import base64
import hashlib
import re
import six
import json
import traceback
import contextlib
from io import BytesIO, StringIO, TextIOWrapper
import requests
from http.client import HTTPConnection  # py3
try:
	from lxml import etree
except Exception:
	pass


from ..websocket._core import WebSocket
from ..env import DEBUG_PRINT_LEVEL
from ..utils.xss_html import XssHtml
from ..utils import events
from ..logging import LOG_WARN, LOG_ERROR, LOG_DEBUG, log_ctx
# CUSTOM IMPORTS
try:
	import openpyxl
	import xlrd
except Exception:
	pass


LOCAL_TZ_TIMEDELTA = datetime.now(timezone.utc)\
	.astimezone().tzinfo.utcoffset(datetime.now())

os.environ['TZ'] = 'UTC'
time.tzset()

# some useful constants
INT64_MAX = 9223372036854775807
MILLIS_IN_MINUTE = 60 * 1000
MILLIS_IN_HOUR = 60 * MILLIS_IN_MINUTE
MILLIS_IN_DAY = 24 * MILLIS_IN_HOUR

SECONDS_IN_MINUTE = 60
SECONDS_IN_HOUR = 60 * SECONDS_IN_MINUTE
SECONDS_IN_DAY = 24 * SECONDS_IN_HOUR

EPOCH = datetime.utcfromtimestamp(0)


_OBJ_END_ = object()
_1KB_ = 1024
_16KB_ = _1KB_ * 16
_32KB_ = _1KB_ * 32
_64KB_ = _1KB_ * 64
_128KB_ = _1KB_ * 128
_1MB_ = _1KB_ * _1KB_


def cur_ms():
	return int(1000 * time.time())


# genetic LRU cache
class LRUCache:
	cache = None
	capacity = None

	def __init__(self, capacity, items=None):
		self.capacity = capacity
		self.cache = collections.OrderedDict(items or [])
		# in addition to recentness of use

	def exists(self, key):
		return self.cache.get(key, None)

	def get(self, key, default=None):
		try:
			# remove and reinsert into
			# ordered dict to move to recent
			value = self.cache.pop(key)
			self.cache[key] = value
			return value
		except KeyError:
			return default

	def set(self, key, value):
		removed_entries = []
		try:
			self.cache.pop(key)
		except KeyError:
			# new entry so cleanup space if it's beyond capacity
			while(len(self.cache) >= self.capacity):
				removed_entries.append(self.cache.popitem(last=False))
		self.cache[key] = value

		return removed_entries

	def __setitem__(self, key, value):
		return self.set(key, value)

	def __getitem__(self, key, default=None):
		return self.get(key, default)

	def __delitem__(self, key):
		return self.cache.pop(key, None)

	def delete(self, key):
		return self.cache.pop(key, None)

	def clear(self):
		return self.cache.clear()

	def to_son(self):
		ret = {}
		for key, val in self.cache:
			ret[key] = val.to_son() if hasattr(val, "to_son") else val
		return ret


class ExpiringCache:
	cache = None
	capacity = None
	ttl = None

	# ttl in millis, 5 minutes default
	def __init__(self, capacity, items=None, ttl=3 * 60 * 1000):
		self.capacity = capacity
		self.cache = collections.OrderedDict()
		self.ttl = ttl
		if(items):
			for k, v in items:
				self.set(k, v)

	def get(self, key, default=None):
		timestamp_and_value = self.cache.get(key, _OBJ_END_)
		if(timestamp_and_value is _OBJ_END_):
			return default
		if(timestamp_and_value[0] > cur_ms()):
			return timestamp_and_value[1]
		self.cache.pop(key, None)  # expired object
		return default

	def set(self, key, value):
		removed_entries = []
		cur_timestamp = cur_ms()
		while(len(self.cache) >= self.capacity):
			removed_entries.append(self.cache.popitem(last=False))
		keys_to_remove = []
		for _key, val in self.cache.items():
			if(val[0] > cur_timestamp):
				break
			# remove expired keys
			keys_to_remove.append(_key)

		for _key in keys_to_remove:
			removed_entries.append(self.cache.pop(_key))

		self.cache[key] = (cur_timestamp + self.ttl, value)
		return removed_entries

	def __setitem__(self, key, value):
		return self.set(key, value)

	def __getitem__(self, key, default=None):
		return self.get(key, default)

	def __delitem__(self, key):
		return self.cache.pop(key, None)

	def exists(self, key):
		return key in self.cache

	def delete(self, key):
		return self.cache.pop(key, None)

	def clear(self):
		return self.cache.clear()

	def to_son(self):
		ret = {}
		cur_timestamp = cur_ms()
		for key, _val in self.cache.items():
			expires_at, val = _val
			if(cur_timestamp < expires_at):
				ret[key] = val.to_son() if hasattr(val, "to_son") else val
		return ret


# get SON of the fields from any generic python object
def to_son(obj):
	ret = obj.__dict__
	for k in ret.keys():
		if(ret[k] is None):
			del ret[k]
	return ret


def from_kwargs(cls, **kwargs):
	ret = cls()
	for key in kwargs:
		setattr(ret, key, kwargs[key])
	return ret


def get_by_key_list(d, key_list, default=None):
	try:
		for key in key_list:
			d = d[key]
		return d
	except (KeyError, IndexError, TypeError):
		return default


def set_by_key_list(d, key_list, value):
	if(not key_list):
		return
	for key in key_list[0:-1]:
		if(key not in d):
			d[key] = {}
		d = d[key]
	d[key_list[-1]] = value


def get_by_key_path(d, key_path, i=0, path="", include_path=False):
	if(not key_path):
		return d
	if(isinstance(key_path, str)):
		key_path = key_path.replace("[", ".[").replace("..", ".").split(".")

	if(i >= len(key_path)):
		return d if not include_path else (d, path)

	key = key_path[i]
	if(key[0] == "["):
		if(isinstance(d, (list, tuple))):
			specific_indexes = [int(x) for x in key[1:-1].split(",") if x]
			ret = []
			if(not specific_indexes):   # just .[].xxx -> all elements
				for _i, val in enumerate(d):
					_ret = get_by_key_path(val, key_path, i + 1, path + f".{_i}", include_path)
					if(_ret is not None):
						ret.append(_ret)
			elif(len(specific_indexes) == 1):  # just .[1].xxx -> specific element
				if(specific_indexes[0] < len(d)):
					_i = specific_indexes[0]
					return get_by_key_path(d[_i], key_path, i + 1, path + f".{_i}", include_path)
			else:  # .[1,2,3].xxx -> multiple elements
				for _i in specific_indexes:
					if(_i < len(d)):
						_ret = get_by_key_path(d[_i], key_path, i + 1, path + f".{_i}", include_path)
						if(_ret is not None):
							ret.append(_ret)
			return ret
		return
	elif(isinstance(d, dict)):
		if("," in key):
			ret = {}
			for _key in key.split(","):
				_d = d
				_key_list = _key.split(":")
				exists = True
				for __key in _key_list:
					if(not (_d := _d.get(__key))):
						exists = False
						break
				if(exists):
					_ret = get_by_key_path(_d, key_path, i + 1, path + f".{_key}", include_path)
					if(_ret is not None):
						ret[_key_list[-1]] = _ret
			return ret
		else:
			return get_by_key_path(d.get(key), key_path, i + 1, path + f".{key}", include_path)
	elif(isinstance(d, list)):
		key = int(key)
		if(key < len(d)):
			return get_by_key_path(d[int(key)], key_path, i + 1, path + f".{key}", include_path)
	return


def date2string(date):
	return date.isoformat()


def date2timestamp(dt):
	# datetime to timestamp
	if not isinstance(dt, datetime):
		return dt
	if(six.PY34):
		return dt.replace(tzinfo=timezone.utc).timestamp()
	else:
		timestamp = time.mktime(dt.timetuple()) + dt.microsecond / 1e6
		return timestamp


def timestamp2date(timestamp):
	if not isinstance(timestamp, (int, float)):
		return timestamp
	date = datetime.utcfromtimestamp(timestamp)
	return date


# 1h2m3s
def duration2string(seconds):
	remaining = seconds

	d = remaining // SECONDS_IN_DAY
	remaining = seconds - d * SECONDS_IN_DAY

	h = remaining // SECONDS_IN_HOUR
	remaining = remaining - h * SECONDS_IN_HOUR

	m = remaining // 60
	remaining = remaining - m * 60

	s = remaining

	ret = ""
	if(d):
		ret += f"{d}d"
	if(h):
		ret += f"{h}h"
	if(m):
		ret += f"{m}m"
	if(s):
		ret += f"{s}s"

	return ret


def string2duration(duration_str):
	ret = 0
	splits = re.split(r'([a-zA-Z]+)', duration_str)
	for i in range(0, len(splits) - len(splits) % 2, 2):
		v = int(splits[i])
		c = splits[i + 1]
		if(c[0] == 'd'):
			ret += v * SECONDS_IN_DAY
		if(c[0] == 'h'):
			ret += v * SECONDS_IN_HOUR
		if(c[0] == 'm'):
			ret += v * 60
		if(c[0] == 's'):
			ret += v
	return ret


def zrfill(tup, n):
	return tup + tuple(0 for x in range(n - len(tup)))


def zlfill(tup, n):
	return tuple(0 for x in range(n - len(tup))) + tup


# remove an item from a set
# return True if item was removed
# return False if item was not present
def set_pop(s, item):
	try:
		s.remove(item)
		return True
	except KeyError:
		return False


DATE_DD_MM_YYYY_REGEX = re.compile(r"(?:^|\s)(\d{1,2})(?:th|nd|rd|st)?[\/\-\.\s\,]+(\d{1,2}|(?:(?:jan|feb|mar|apr|may|june|jul|aug|sep|oct|nov|dec)\w*))(?:[\/\-\.\s](\d{4}))?(?=\s|$)", re.IGNORECASE)
TIME_REGEX = re.compile(r"(\d{1,2})(?:\s*:+(\d{1,2}))?(?:\s*:+(\d{1,2}))?(?::+(\d{1,2})|(\s*(?:a|p).?m))", re.IGNORECASE)
DAY_REGEX = re.compile(r"(mon|tues|wed(nes)?|thur(s)?|fri|sat(ur)?|sun)(?!t)", re.IGNORECASE)
RELATIVE_DAY_REGEX = re.compile(r"(after\s+to|tod|tom)[^\s]*", re.IGNORECASE)

DAYS_OF_WEEK = [
	'monday', 'tuesday', 'wednesday',
	'thursday', 'friday', 'saturday', 'sunday'
]
MONTHS_OF_YEAR = {
	'jan': 1, 'feb': 2, 'mar': 3,
	'apr': 4, 'may': 5, 'jun': 6,
	'jul': 7, 'aug': 8, 'sep': 9,
	'oct': 10, 'nov': 11, 'dec': 12
}


def get_overlap(start, end, a, b, partial):
	if(partial):
		if(start >= a and start <= b):
			return (start, (end if end < b else b))
		elif(start <= a and a <= end):
			return (a, (end if end < b else b))
	else:
		if(start >= a and end <= b):
			return (start, end)


# tz_delta = local - UTC ( to dereive UTC times in current timezone)

def _iter_time_overlaps(a, b, x: str, tz_delta, partial=False, interval=None):
	x, *params = x.split("|")
	x = x.lower().strip()
	if(isinstance(a, int)):
		if(a > 1e11):
			a = a // 1000
		a = timestamp2date(a)

	if(isinstance(b, int)):
		if(b > 1e11):
			b = b // 1000
		b = timestamp2date(b)
	x = x.strip()

	if("weekend" in x):
		x = x.replace("weekend", "sat,sun")

	if("next" in x):  # some keyword checks
		if("month" in x):
			_month = a.month + 1
			a = datetime(a.year + _month // 12, _month % 12, 1)
			x = x.replace("month", "")
		elif("week" in x):
			a = datetime(a.year, a.month, a.day) + timedelta(days=7 - a.weekday())
			x = x.replace("week", "")
		else:
			a = datetime(a.year, a.month, a.day) + timedelta(days=1)
			# next day
			x = x.replace("day", "")
		x = x.replace("next", "")

	time_matches = []
	date_matches = []
	day_matches = []

	time_match_pos_list = []
	date_match_pos_list = []
	day_match_pos_list = []

	# find all time matches
	for match in TIME_REGEX.finditer(x):
		h, *ms, am_pm = match.groups()
		ms = [x for x in ms if x is not None] + [None, None]
		m, s = ms[:2]
		time_matches.append([int(h or 0), int(m or 0), int(s or 0), (am_pm or "").strip().lower()])
		time_match_pos_list.append(match.span())

	# RELATIVE DAY MATCHES
	dt_now = datetime.now()
	for match in RELATIVE_DAY_REGEX.finditer(x):
		day_str = match.group().lower()
		if(day_str.startswith("tod")):
			date_matches.append([dt_now.day, dt_now.month, dt_now.year])
			date_match_pos_list.append(match.span())
		elif(day_str.startswith("tom")):
			dt_tommorow = dt_now + timedelta(days=1)
			date_matches.append([dt_tommorow.day, dt_tommorow.month, dt_tommorow.year])
			date_match_pos_list.append(match.span())
		elif(day_str.startswith("after")):
			# day fter tomorrow
			dt_day_after_tom = dt_now + timedelta(days=2)
			date_matches.append([dt_day_after_tom.day, dt_day_after_tom.month, dt_day_after_tom.year])
			date_match_pos_list.append(match.span())

	# date matches
	_dt_a_parts = [a.day, a.month, a.year]
	for match in DATE_DD_MM_YYYY_REGEX.finditer(x):
		_dt_parts = [_x for _x in match.groups() if _x is not None]
		dt_parts = _dt_parts + _dt_a_parts[len(_dt_parts):]
		if(month := MONTHS_OF_YEAR.get(dt_parts[1][:3].lower())):
			dt_parts[1] = month
		date_matches.append([int(_x) for _x in dt_parts])
		date_match_pos_list.append(match.span())

	# day matches
	for match in DAY_REGEX.finditer(x):
		day_str = x[match.start():match.end()].lower()
		day_index = next((i for i, day in enumerate(DAYS_OF_WEEK) if day.startswith(day_str)), None)
		if(day_index is not None):
			day_matches.append(day_index)
			day_match_pos_list.append(match.span())

	# is it date-date | time-time or date time - date time
	is_time_per_day = bool(time_match_pos_list and (
		(
			date_match_pos_list
			and not get_overlap(
				date_match_pos_list[0][0], date_match_pos_list[-1][0],
				time_match_pos_list[0][0], time_match_pos_list[-1][0],
				True
			)
		) or (
			day_match_pos_list
			and not get_overlap(
				day_match_pos_list[0][0], day_match_pos_list[-1][0],
				time_match_pos_list[0][0], time_match_pos_list[-1][0],
				True
			)
		)
	))

	for time_match in time_matches:
		if(time_match[-1] and time_match[-1][0] in ("a", "p")):
			while(len(time_match) < 4):
				time_match.insert(-1, 0)
		else:
			while(len(time_match) < 4):
				time_match.append(0)

		for i in range(3):
			time_match[i] = int(time_match[i])

	# offset from start of day
	offset_check_x = timedelta()
	offset_check_y = timedelta(hours=24)
	if(time_matches):
		hours, mins, secs, am_pm = time_matches[0]
		if(am_pm):
			# convert to 24 hour format
			if(am_pm.lower()[0] == "p"):
				hours = (hours % 12) + 12
			if(am_pm.lower()[0] == "a"):
				if(hours == 12):
					hours = 0
		offset_check_x = timedelta(hours=hours, minutes=mins, seconds=secs)

		if(len(time_matches) > 1):
			hours, mins, secs, am_pm = time_matches[1]

			if(am_pm):
				# convert to 24 hour format
				if(am_pm.lower()[0] == "p"):
					hours = (hours % 12) + 12
				if(am_pm.lower()[0] == "a"):
					if(hours == 12):
						hours = 0
			offset_check_y = timedelta(hours=hours, minutes=mins, seconds=secs)
		else:
			offset_check_y = offset_check_x + timedelta(seconds=interval or SECONDS_IN_HOUR)  # default 1 hour

		if(offset_check_y <= offset_check_x):
			offset_check_y += timedelta(days=1)

	if(date_matches):
		day, month, year, *_ = map(lambda x: int(x) if x else 0, date_matches[0])
		if(year < 100):
			year = 2000 + year
		date_range_start = datetime(year=year, month=month, day=day)
		if(len(date_matches) > 1):
			day, month, year, *_ = map(lambda x: int(x) if x else 0, date_matches[1])
			if(year < 100):
				year = 2000 + year
		date_range_end = datetime(year=year, month=month, day=day)

		if(is_time_per_day):  # date-date time-time
			while(date_range_start <= date_range_end):
				if(_ret := get_overlap(
					date_range_start + offset_check_x - tz_delta,
					date_range_start + offset_check_y - tz_delta,
					a, b, partial
				)):
					yield *_ret, params
				date_range_start += timedelta(days=1)
		else:  # date time - date time
			_ret = get_overlap(
				date_range_start + offset_check_x - tz_delta,
				date_range_end + offset_check_y - tz_delta,
				a, b, partial
			)
			if(_ret):
				yield *_ret, params

	elif(day_matches):
		start_day = end_day = day_matches[0]
		if(len(day_matches) > 1):
			end_day = day_matches[1]

		segments = [(start_day, end_day)]\
			if(end_day >= start_day) \
			else [(start_day, 6), (0, end_day)]

		# should be atleast one match
		found = False
		for i in range(7):
			t = a + timedelta(days=i)
			t_weekday = t.weekday()
			for start_day, end_day in segments:
				if(start_day <= t_weekday <= end_day):
					# found a day in the week that matches
					found = True
					break
			if(found):
				break

		t_start_of_day = datetime(year=t.year, month=t.month, day=t.day)
		if(not is_time_per_day):
			# find the day from now that matches start_day
			if(_ret := get_overlap(
				t_start_of_day + offset_check_x - tz_delta,
				t_start_of_day + timedelta(days=abs(end_day - start_day))
					+ offset_check_y - tz_delta,
				a, b, partial
			)):
				yield *_ret, params
		else:
			while(t_start_of_day <= b):
				for i in range(abs(end_day - start_day) + 1):
					if((t2_start_of_day := t_start_of_day + timedelta(days=i)) <= b):
						if(_ret := get_overlap(
							t2_start_of_day + offset_check_x - tz_delta,
							t2_start_of_day + offset_check_y - tz_delta,
							a, b, partial
						)):
							yield *_ret, params
				t_start_of_day += timedelta(days=7)

	elif(time_matches):  # ONLY TIMES
		t_start_of_day = datetime(a.year, a.month, a.day)
		while(t_start_of_day < b):
			if(_ret := get_overlap(
				t_start_of_day + offset_check_x - tz_delta,
				t_start_of_day + offset_check_y - tz_delta,
				a, b, partial
			)):
				yield *_ret, params
			t_start_of_day += timedelta(days=1)

	return []


def iter_time_overlaps(
	a, b, include: list, exclude=None,
	partial=False, tz_delta=timedelta(),
	milliseconds=False, interval=None, delim=","
):
	if(isinstance(include, str)):
		include = include.split(delim)
	if(isinstance(exclude, str)):
		exclude = exclude.split(delim)

	if(
		interval is not None
		and isinstance(interval, str)
	):
		interval = string2duration(interval)

	buffer = []
	heapq.heapify(buffer)
	for i, x in enumerate(include):  # i -> tie breaker
		try:
			it = _iter_time_overlaps(
				a, b, x, tz_delta, partial=partial,
				interval=interval
			)
			heapq.heappush(buffer, (next(it), x, i, it))
		except StopIteration:
			pass

	while(len(buffer) > 0):
		time_range, x, i, it = heapq.heappop(buffer)  # pop the least i -> tie breaker
		try:
			heapq.heappush(buffer, (next(it), x, i + 1, it))
		except StopIteration:
			pass
		# check if it's excluded
		if(
			not exclude
			or not get_time_overlaps(
				time_range[0], time_range[1], exclude,
				limit=1, partial=True, tz_delta=tz_delta
			)
		):
			if(interval is not None):
				start, end, params = time_range
				while((_end := start + timedelta(seconds=interval)) <= end):
					if(milliseconds):
						yield (
							int(date2timestamp(start) * 1000),
							int(date2timestamp(_end) * 1000),
							params
						)
					else:
						yield (start, _end, params)
					start = _end
			else:
				if(milliseconds):
					yield (
						int(date2timestamp(time_range[0]) * 1000),
						int(date2timestamp(time_range[1]) * 1000),
						*time_range[2:]
					)
				else:
					yield time_range
	return


def get_time_overlaps(
	a, b, include: list, exclude=None,
	partial=False, tz_delta=timedelta(), milliseconds=False,
	limit=10, interval=None, delim=","
):
	iter = iter_time_overlaps(
		a, b, include, exclude=exclude, partial=partial,
		tz_delta=tz_delta, milliseconds=milliseconds,
		interval=interval, delim=delim
	)
	ret = []
	for i in range(limit):
		time_range = next(iter, None)
		if(not time_range):
			break
		ret.append(time_range)
	return ret


def get_start_of_day(t):
	return datetime(year=t.year, month=t.month, day=t.day)


def get_start_of_day_millis(millis):
	return (millis // MILLIS_IN_DAY) * MILLIS_IN_DAY


def find_index(a_list, value, start=0):
	try:
		return a_list.index(value, start)
	except ValueError:
		return -1


def find_nth(haystack, needle, n):
	''' Find position of nth occurance in a string'''
	if(not haystack):
		return -1
	start = find_index(haystack, needle)
	while start >= 0 and n > 1:
		start = find_index(haystack, needle, start + len(needle))
		n -= 1
	return start


SOME_TIME_WHEN_WE_STARTED_THIS = 1471504855
SOME_TIME_WHEN_WE_STARTED_THIS_MILLIS_WITH_10 = 16111506425808


# this function is inspired from instagram engineering
# post on generating 64bit keys with 12 bit shard_id inside it
# __thread_data = LRUCache(10000)
def generate_64bit_key(shard_id):  # shard id should be a 12 bit number
	# may be later use dattime
	millis_elapsed = int((time.time() - SOME_TIME_WHEN_WE_STARTED_THIS) * 1000)

	# 41 bits , clear 22 places for random id and shard_id
	_id = (((1 << 41) - 1) & millis_elapsed) << 23

	# 12 bit shard id, on top
	_id |= (((1 << 12) - 1) & shard_id) << 11

	# increment per thread number
	# thread_id = threading.current_thread().ident
	# thread_data = __thread_data.get(thread_id, None)
	# if(not thread_data):
	#     thread_data = [0]
	#     __thread_data.set(thread_id , thread_data)
	# thread_data[0] = (thread_data[0] + 1)%(1 << 11)
	random_11_bits = random.randrange(0, 1 << 11)
	_id |= (((1 << 11) - 1) & random_11_bits)  # clear 12 places for random

	# shard_id|timestmap|random_number

	return _id


def int_to_bytes(number: int) -> bytes:
	return number.to_bytes(
		length=(8 + (number + (number < 0)).bit_length()) // 8,
		byteorder='big',
		signed=True
	)


BASE_62_LIST = string.ascii_uppercase + string.digits + string.ascii_lowercase
BASE_62_DICT = dict((c, i) for i, c in enumerate(BASE_62_LIST))


def str_to_int(string, reverse_base=BASE_62_DICT):
	length = len(reverse_base)
	ret = 0
	for i, c in enumerate(string[::-1]):
		ret += (length ** i) * reverse_base[c]

	return ret


def int_to_str(integer, base=BASE_62_LIST):
	if integer == 0:
		return base[0]

	length = len(base)
	ret = ''
	while integer != 0:
		ret = base[integer % length] + ret
		integer //= length

	return ret


_pev_int_id = 0


def get_int_id():
	time_elapsed \
		= int(time.time() * 10000 - SOME_TIME_WHEN_WE_STARTED_THIS_MILLIS_WITH_10)

	global _pev_int_id
	if(time_elapsed <= _pev_int_id):
		time_elapsed = _pev_int_id + 1
	_pev_int_id = time_elapsed
	return time_elapsed


def get_str_id():
	return int_to_str(get_int_id())


def is_almost_equal(a, b, max_diff=0.01):
	if(a is None and b is None):
		return True
	if(a is None or b is None):
		return False
	diff = abs(a - b)
	if(diff < max_diff):
		return False
	return True


# protobuf + mysql utility functions
def list_to_protobuf(values, message):
	'''parse list to protobuf message'''
	if not values:
		return
	if isinstance(values[0], dict):  # value needs to be further parsed
		for v in values:
			cmd = message.add()
			dict_to_protobuf(v, cmd)
	else:  # value can be set
		message.extend(values)


# just like move constructor ;) , move values much faster than copies
def dict_to_protobuf(
	values, obj,
	transformations=None, excludes=None, preserve_values=True
):
	if(not preserve_values):
		if(transformations):
			for k, func in transformations.items():
				values[k] = func(values.get(k, None))

		if(excludes):
			for exclude, flag in excludes.items():
				if(hasattr(values, exclude)):
					del values[exclude]

	for k, v in values.items():
		if(preserve_values):
			if(transformations and k in transformations):
				v = transformations[k](v)

			if(excludes and k in excludes):
				continue

		if hasattr(obj, k):
			if isinstance(v, dict):  # value needs to be further parsed
				dict_to_protobuf(v, getattr(obj, k))
			elif isinstance(v, list):
				list_to_protobuf(v, getattr(obj, k))
			else:  # value can be set
				if v:  # otherwise default
					setattr(obj, k, v)


def get_mysql_rows_as_dict(res):
	rows_as_dict = []
	for row in res.rows:
		as_dict = {}
		for field, val in zip(res.fields, row):
			as_dict[field[0]] = val

		rows_as_dict.append(as_dict)

	return rows_as_dict

########################


def remove_duplicates(lst, key=None):
	exists = set()
	ret = []
	for i in lst:
		if(key):
			if((key_val := key(i)) in exists):
				continue
			exists.add(key_val)
		else:
			if(i in exists):
				continue
			exists.add(i)
		# i doesn't exist
		ret.append(i)
	return ret


def get_empty_fields(_dict, *fields):
	ret = []
	for field in (fields or _dict.keys()):
		if(not _dict.get(field, None)):
			ret.append(field)
	return ret


def get_shard_id(_id):
	return (int(_id) & (((1 << 12) - 1) << 11)) >> 11


# can split at most at n points
# always returns n + 1 parts
def nsplit(_str, delimiter, n):
	parts = _str.split(delimiter, n)
	for i in range(n + 1 - len(parts)):
		parts.append(None)
	return parts


def utf8(value) -> bytes:
	if(isinstance(value, bytes)):
		return value
	return value.encode("utf-8")


def hmac_hexdigest(secret, *parts) -> str:
	hash = hmac.new(utf8(secret), digestmod=hashlib.sha256)
	for part in parts:
		hash.update(utf8(part))
	return hash.hexdigest()


def create_signed_value(name, value: bytes | str, secret, expires_in=31 * SECONDS_IN_DAY, values=()) -> str:
	expires_at = str(int(time.time() + expires_in))
	value = base64.b64encode(utf8(value)).decode()  # hide value
	# ~ to prevent changing value, timestamp but value + timestamp being same
	signature = hmac_hexdigest(secret, name, value, *values, expires_at)
	signed_value = "|".join([value, *values, expires_at, signature])
	return signed_value


def decode_signed_value(name, signed_value, secret, expiry_check=True) -> bytes:
	if not signed_value:
		return None
	_parts = signed_value.split("|")
	if len(_parts) < 3:
		return None

	_value, *_values, expires_at, _signature = _parts
	# check signature matches or not
	signature = hmac_hexdigest(secret, name, _value, *_values, expires_at)
	if _signature != signature:
		return None

	if(expiry_check):
		expires_at = int(expires_at)
		if(time.time() > expires_at):
			return None
	try:
		return base64.b64decode(_value)
	except Exception:
		return None


def dangerously_peek_signed_value(value) -> bytes:
	_parts = utf8(value).split(b"|")
	if len(_parts) < 3:
		return None
	_value, *parts, expires_at, _signature = _parts
	try:
		return base64.b64decode(_value)
	except Exception:
		return None


"""Dependencies are expressed as a dictionary whose keys are items
	and whose values are a set of dependent items. Output is a list of
	sets in topological order. The first set consists of items with no
	dependences, each subsequent set consists of items that depend upon
	items in the preceeding sets.
"""


def toposort(data):

	# Special case empty input.
	if len(data) == 0:
		return

	# Copy the input so as to leave it unmodified.
	data = data.copy()

	# Ignore self dependencies.
	for k, v in data.items():
		v.discard(k)
	# Find all items that don't depend on anything.
	extra_items_in_deps = _reduce(set.union, data.values()) - set(data.keys())
	# Add empty dependences where needed.
	data.update({item: set() for item in extra_items_in_deps})
	while True:
		ordered = set(item for item, dep in data.items() if len(dep) == 0)
		if not ordered:
			break
		yield ordered
		data = {
			item: (dep - ordered)
			for item, dep in data.items() if item not in ordered
		}
	if len(data) != 0:
		exception_string = \
			'Circular dependencies exist among these items: {{{}}}'.format(
				', '.join([
					'{!r}:{!r}'.format(key, value) for key, value in sorted(data.items())
				])
			)
		raise Exception(exception_string)


# Console or Cloud Console.
def get_random_id(length=10, include_timestamp=True):
	'''returns a random string containing numbers lowercase upper case'''
	'''http://stackoverflow.com/questions/2257441/random-string-generation-with-upper-case-letters-and-digits-in-python'''

	key_str = ''.join(
		random.SystemRandom().choice(BASE_62_LIST) for _ in range(length)  # iter
	)
	if(include_timestamp):
		key_str = f"{int(time.time())}{key_str}"
	# key_str = hashlib.md5(key_str).hexdigest()
	return key_str


# jump consistent hash function , jump_hash("Asdasd", 100) assigns to a bucket
def jump_hash(key, num_buckets):
	b, j = -1, 0
	key = int(hashlib.md5(key).hexdigest(), 16)
	while j < num_buckets:
		b = int(j)
		key = ((key * int(2862933555777941757)) + 1) & 0xFFFFFFFFFFFFFFFF
		j = float(b + 1) * (float(1 << 31) / float((key >> 33) + 1))
	return int(b)

##############


INT_REGEX = re.compile(r"(-?[0-9]+)")
NUMBER_REGEX = re.compile(r"(-?[0-9\.]+)")
NON_ALPHA_NUM_SPACE_DOT_REGEX = re.compile(r"[^0-9a-zA-Z \.]", re.DOTALL)  # space, . allowed
NON_ALPHA_REGEX = re.compile(r"[^0-9a-zA-Z]", re.DOTALL)
NON_ALPHA_NUM_UNDERSCORE_REGEX = re.compile(r"[^0-9a-zA-Z_]", re.DOTALL)
NON_ALPHA_GROUPS_REGEX = re.compile(r"[^0-9a-zA-Z]+", re.DOTALL)  # multiple non-alpha groups


def sanitize_string(text):
	return NON_ALPHA_NUM_SPACE_DOT_REGEX.sub("", text)


def sanitize_to_id(text):
	return NON_ALPHA_REGEX.sub("", text.lower())


def sanitize_to_id_allow_case(text):
	return NON_ALPHA_REGEX.sub("", text)


def sanitize_to_underscore_id(text):
	return NON_ALPHA_GROUPS_REGEX.sub("_", text.strip().lower()).rstrip("_")


def sanitize_to_underscore_id_allow_case(text):
	return NON_ALPHA_GROUPS_REGEX.sub("_", text.strip()).rstrip("_")


EMAIL_REGEX = re.compile(
	r'^[_a-z0-9-]+(\.[_a-z0-9-]+)*(\+[_a-z0-9-]+)?\@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,6})$',
	re.IGNORECASE
)


def is_valid_email(email):
	return EMAIL_REGEX.match(email)


SANITIZE_EMAIL_REGEX_NO_PLUS_ALLOWED = re.compile(r"\+[^@]*")


def sanitize_email_id(email_id, plus_allowed=True):
	if(not email_id):
		return None
	email_id = email_id.strip().lower()
	if(not EMAIL_REGEX.match(email_id)):
		return None

	if(not plus_allowed):
		email_id = SANITIZE_EMAIL_REGEX_NO_PLUS_ALLOWED.sub("", email_id)

	return email_id


# first minus second
def list_diff(first, second):
	if(not first or not second):
		return first
	second = set(second)
	return [item for item in first if item not in second]


# what are added vs what need to be deleted from first
# no order
def list_diff2(first, second, key=None):
	if(not first and second):
		return [], list(second)
	if(second is None):
		return list(first), []
	if(key):
		second_keys = set(map(key, second))
		first_keys = set(map(key, first))
		# to add and delete
		return (
			[item for item in first if key(item) not in second_keys],
			[item for item in second if key(item) not in first_keys]
		)
	else:
		second = set(second)
		first = set(first)
		# to add and delete
		return (
			[item for item in first if item not in second],
			[item for item in second if item not in first]
		)


def compress_lists(lists) -> dict:
	'''returns a tree like structure from a list of lists'''
	ret = {}
	for lst in lists:
		_next = ret
		for i in lst:
			if(i not in _next):
				_next[i] = {}
			_next = _next[i]
	return ret


def last(arr, default=None):
	return arr[-1] if len(arr) > 0 else default


def first(arr, default=None):
	return arr[0] if len(arr) > 0 else default


# a dummy object with given keys,values
class DummyObject:
	entries = None

	def __init__(self, entries=None, **kwargs):
		self.entries = entries if entries is not None else {}
		if(kwargs):
			self.entries.update(kwargs)

	def __setattr__(self, key, val):
		if(self.entries):
			self.entries[key] = val
		else:
			super().__setattr__(key, val)

	def __getattr__(self, key):
		return self.entries.get(key)

	def get(self, key, default=None):
		return self.entries.get(key, default)

	def __getitem__(self, key):
		return self.entries.get(key)

	def __setitem__(self, key, val):
		self.entries[key] = val

	def __repr__(self):
		return str(self.entries)

	def to_dict(self):
		return self.entries


# fully qualified name of the object
def object_name(o):
	# o.__module__ + "." + o.__class__.__qualname__ is an example in
	# this context of H.L. Mencken's "neat, plausible, and wrong."
	# Python makes no guarantees as to whether the __module__ special
	# attribute is defined, so we take a more circumspect approach.
	# Alas, the module name is explicitly excluded from __qualname__
	# in Python 3.

	module = o.__class__.__module__
	if module is None or module == str.__class__.__module__:
		return o.__class__.__name__  # Avoid reporting __builtin__
	else:
		return module + '.' + o.__class__.__name__


# normalize a int/str to n number of characters
def normalize(val, n=13):
	val = str(val)
	length = len(val)
	if(n >= length):
		return ("0" * (n - length)) + val

	raise Exception('cannot have greater length, ensure about this')


# normalize a int/str to n number of characters
# and clips of any characters more than n characters
def normalize_no_warn(val, n=13):
	val = str(val)
	length = len(val)
	if(n >= length):
		return ("0" * (n - length)) + val
	return val


def ltrim(_str, str_to_remove):
	if(str_to_remove and _str.startswith(str_to_remove)):
		return _str[len(str_to_remove):]
	return _str


def rtrim(_str, str_to_remove):
	if(str_to_remove and _str.endswith(str_to_remove)):
		return _str[:len(_str) - len(str_to_remove)]
	return _str


def trim(_str, str_to_remove):
	return rtrim(ltrim(_str, str_to_remove), str_to_remove)


# remove zeros in the front
# returns 0 if it's all zeroes
def de_normalize(id1):
	i = 0
	length = len(id1)
	while(i < length):
		if(id1[i] != '0'):
			break
		i = i + 1
	if(i >= length):
		return '0'
	return id1[i:]


def set_non_blocking(fd):
	"""
	Set the file description of the given file descriptor to non-blocking.
	"""
	flags = fcntl.fcntl(fd, fcntl.F_GETFL)
	flags = flags | os.O_NONBLOCK
	fcntl.fcntl(fd, fcntl.F_SETFL, flags)


class ThreadPool:
	def worker(self):
		while((val := self.tasks.get()) is not None):
			# get a task from queue
			func, args, kargs = val
			try:
				func(*args, **kargs)
			except Exception as ex:
				# An exception happened in this thread
				LOG_ERROR(
					f"threadpool:{self.name}",
					exception_str=str(ex),
					stacktrace_string=traceback.format_exc(),
					args=str(args),
					kwargs=str(kargs)
				)

	def __init__(self, num_threads, name="default"):
		self.tasks = Queue(num_threads)
		self.worker_threads = []
		self.name = name
		for _ in range(num_threads):
			thread = Thread(target=self.worker)
			self.worker_threads.append(thread)
			thread.start()

	def add_task(self, func, *args, **kargs):
		""" Add a task to the queue """
		self.tasks.put((func, args, kargs))

	def stop(self):
		""" Wait for completion of all the tasks in the queue """
		for _ in range(len(self.worker_threads)):
			self.tasks.put(None)  # cleanup all threads
		# join all threads
		for worker_thread in self.worker_threads:
			worker_thread.join()


def raise_exception(msg):
	raise Exception(msg)


def make_xss_safe(_html):
	if(not _html):
		return _html
	parser = XssHtml()
	parser.feed(_html)
	parser.close()
	return parser.getHtml()


# yeild batches of items from iterator
def batched_iter(iterable, n=1):
	current_batch = []
	for item in iterable:
		current_batch.append(item)
		if len(current_batch) == n:
			yield current_batch
			current_batch = []
	if current_batch:
		yield current_batch


# This is the most *important* socket wrapped implementation
# used by blaster server.
class BufferedSocket():

	is_eof = False
	sock = None
	store = None
	lock = None

	def __init__(self, sock):
		self.sock = sock
		self.readbuf = bytearray()
		self.sendbuf = bytearray()

	def close(self):
		self.sock.close()
		self.is_eof = True

	def sendb(self, *_data):  # send but buffered
		n = 0
		for data in _data:
			if(isinstance(data, str)):
				data = data.encode()
			n += len(data)
			self.sendbuf.extend(data)
		if(len(self.sendbuf) > _1MB_):
			return self.flush()
		return n

	def send(self, *_data):
		for data in _data:
			if(isinstance(data, str)):
				data = data.encode()
			self.sendbuf.extend(data)
		return self.flush()

	def flush(self):
		if(len(self.sendbuf) == 0):
			return 0
		send_buf = self.sendbuf
		self.sendbuf = bytearray()  # reset
		return self.sock.sendall(send_buf)

	def sendl(self, *_data):
		if(not self.lock):
			self.lock = BoundedSemaphore()
		self.lock.acquire()
		ret = self.send(*_data)
		self.lock.release()
		return ret

	def recv(self, n):
		if(len(self.readbuf)):
			ret = self.readbuf
			self.readbuf = bytearray()
			return ret
		return self.sock.recv(n)

	def recvn(self, n):
		while(len(self.readbuf) < n):
			data = self.sock.recv(4096)
			if(not data):
				self.is_eof = True
				return self.readbuf or None
			self.readbuf.extend(data)

		# return n bytes for now
		ret = self.readbuf[:n]
		# set remaining to new store
		self.readbuf = self.readbuf[n:]
		return ret

	# fails if it couldn't find the delimiter until max_size
	def readuntil(self, delimiter, max_size, discard_delimiter):
		# check in store
		if(isinstance(delimiter, str)):
			delimiter = delimiter.encode()

		delimiter_len = len(delimiter)
		# scan the store until end, if not found extend
		# and continue until store > max_size
		to_scan_len = len(self.readbuf)
		i = 0  # how much we scanned already

		_store = self.readbuf  # get a reference
		while(True):
			if(i > max_size):
				self.is_eof = True
				return None
			if(i >= delimiter_len):
				j = 0
				lookup_from = i - delimiter_len
				while(j < delimiter_len and _store[lookup_from + j] == delimiter[j]):
					j += 1

				if(j == delimiter_len):
					# found
					ret = None
					if(discard_delimiter):
						ret = _store[:i - delimiter_len]
					else:
						ret = _store[:i]
					self.readbuf = _store[i:]  # set store to unscanned/pending
					return ret
			if(i >= to_scan_len):
				# scanned all buffer
				data = self.sock.recv(4096)
				if(not data):
					self.is_eof = True
					return None

				_store.extend(data)  # fetch more data
				to_scan_len = len(_store)
			i += 1

	def __getattr__(self, key):
		ret = getattr(self.sock, key, _OBJ_END_)
		if(ret is _OBJ_END_):
			raise AttributeError()
		return ret


# this variable indicated the TCP_USER_TIMEOUT
# parameter that indicated after how long without an
# ack packet we should close
__tcp_user_timeout = 30 * 1000


def set_socket_fast_close_options(sock):
	# abruptly close the connection after 10 seconds
	# without back and forth communication about closing
	# i.e waiting in time_wait state
	sock.setsockopt(
		socket.SOL_SOCKET, socket.SO_LINGER,
		struct.pack('ii', 1, 10)
	)
	# it can wait 10 seconds,
	# if there is congestion on the network card to send data
	sock.setsockopt(
		socket.SOL_SOCKET, socket.SO_SNDTIMEO,
		struct.pack('ll', 10, 0)
	)

	# after 30 seconds if there is no ack
	# then we assume broken and close it
	TCP_USER_TIMEOUT = 18
	sock.setsockopt(socket.SOL_TCP, TCP_USER_TIMEOUT, 30 * 1000)


# wraps send method of websocket which keeps a buffer of messages
# for 20 seconds if the connection closes, you can use them to resend
class WebsocketConnection(WebSocket):
	# to keep track of how many greenlets are waiting on semaphore to send
	queue = None
	# queue for older sent messages in case of reset we try to retransmit
	msg_assumed_sent = None
	ws = None
	lock = BoundedSemaphore()
	is_stale = False
	last_msg_recv_timestamp = None
	last_msg_sent_timestamp = None

	user_id = None
	is_viewer_list = False

	def __init__(self, ws, user_id):
		self.ws = ws
		# to keep track of how many greenlets are waiting on semaphore to send
		self.queue = collections.deque()
		# queue for older sent messages in case of reset we try to retransmit
		self.msg_assumed_sent = collections.deque()
		self.user_id = user_id

		self.last_msg_recv_timestamp\
			= self.last_msg_sent_timestamp \
			= time.time() * 1000

	# msg is only string data , #ref is used
	# just in case an exception occurs , we pass that ref
	def send(self, msg, ref=None, add_to_assumend_sent=True):
		if(self.is_stale):
			raise Exception("stale connection")

		self.queue.append((ref, msg))
		if(self.lock.locked()):
			return

		self.lock.acquire()
		data_ref = None
		data = None
		try:
			while(not self.is_stale and len(self.queue) > 0):
				data_ref, data = self.queue.popleft()  # peek
				self.ws.send(data)  # msg objects only
				current_timestamp = time.time() * 1000
				self.last_msg_sent_timestamp = current_timestamp
				if(add_to_assumend_sent):
					while(
						len(self.msg_assumed_sent) > 0
						and self.msg_assumed_sent[0][0] < current_timestamp - __tcp_user_timeout
					):
						# keep inly 20 seconds of previous data
						self.msg_assumed_sent.popleft()

					self.msg_assumed_sent.append((current_timestamp, data_ref, data))
		except Exception:
			err_msg = "Exception while sending message to {}, might be closed ".format(
				self.user_id
			)
			self.is_stale = True
			raise Exception(err_msg)
		finally:
			self.lock.release()
		return


def parse_cmd_line_arguments():
	from sys import argv
	args = []
	args_map = {}
	i = 0
	num_args = len(argv)
	while(i < num_args):
		arg = argv[i]
		if(arg.startswith("-")):
			if("=" in arg):
				key, val = arg.split("=", 1)
				args_map[key.lstrip("-")] = val
			else:
				next_arg = True
				if(i + 1 < num_args and not argv[i + 1].startswith("-")):
					next_arg = argv[i + 1]
					i += 1
				args_map[arg.lstrip("-")] = next_arg
		else:
			args.append(arg)

		i += 1
	return args, args_map


# PARSE COMMAND LINE ARGUMENTS BY DEFAULT
CommandLineArgs, CommandLineNamedArgs = parse_cmd_line_arguments()


def run_shell(
	cmd, output_parser=None, shell=False, max_buf=5000, fail=True, state=None, env=None,
	process_lines=None, **kwargs
):

	DEBUG_PRINT_LEVEL > 2 and print(f"#RUNNING: {cmd}")
	state = state if state is not None else DummyObject()
	state.total_output = ""
	state.total_err = ""
	last_new_line_indexes = [-1, -1]  # output, err

	# keep parsing output
	def process_output(proc_out, proc_in):
		while(_out := proc_out.read(1)):
			# add to our input
			state.total_output += _out

			if(process_lines is not None and _out == "\n"):
				process_lines(state.total_output[last_new_line_indexes[0] + 1:], None)
				last_new_line_indexes[0] = len(state.total_output) - 1

			if(len(state.total_output) > 2 * max_buf):  # clip output to max_buf
				state.total_output = state.total_output[-max_buf:]
				last_new_line_indexes[0] = max(-1, last_new_line_indexes[0] - max_buf)

			if(output_parser):
				# parse the output and if it returns something
				# we write that to input file(generally stdin)
				_inp = output_parser(state.total_output, state.total_err)
				if(_inp):
					proc_in.write(_inp)
			else:
				print(_out, end="", flush=True)

	def process_error(proc_err, proc_in):
		while(_err := proc_err.read(1)):
			# add to our input
			state.total_err += _err

			if(process_lines is not None and _err == "\n"):
				process_lines(None, state.total_err[last_new_line_indexes[1] + 1:])
				last_new_line_indexes[1] = len(state.total_err) - 1

			if(len(state.total_err) > 2 * max_buf):
				state.total_err = state.total_err[-max_buf:]
				last_new_line_indexes[1] = max(-1, last_new_line_indexes[1] - max_buf)

			if(output_parser):
				# parse the output and if it returns something
				# we write that to input file(generally stdin)
				_inp = output_parser(state.total_output, state.total_err)
				if(_inp):
					proc_in.write(_inp)
			else:
				print(_err, end="", flush=True)

	if(isinstance(cmd, str) and not shell):
		cmd = shlex.split(cmd)

	dup_stdin = os.dup(sys.stdin.fileno()) if shell else subprocess.PIPE
	_env = os.environ.copy()
	if(env):
		_env.update(env)

	state.proc = proc = subprocess.Popen(
		cmd,
		stdin=dup_stdin,
		stdout=subprocess.PIPE,
		stderr=subprocess.PIPE,
		shell=shell,
		env=_env,
		**kwargs
	)

	# process output reader
	output_parser_thread = Thread(
		target=process_output,
		args=(
			TextIOWrapper(proc.stdout, encoding="utf-8"),
			proc.stdin
		)
	)
	# process err reader
	err_parser_thread = Thread(
		target=process_error,
		args=(
			TextIOWrapper(proc.stderr, encoding="utf-8"),
			proc.stdin
		)
	)
	output_parser_thread.start()
	err_parser_thread.start()

	os.close(dup_stdin) if dup_stdin != subprocess.PIPE else None

	# just keep printing error
	# wait for process to terminate
	ret_code = proc.wait()
	state.return_code = ret_code
	output_parser_thread.join()
	err_parser_thread.join()
	state.proc = None
	if(ret_code and fail):
		raise Exception(f"Non zero return code :{ret_code}")
	return state


# args is array of strings or array or array of words
# you can use this to return a bunch of strings to index
# in elasticsearch with key "search_words"
def condense_for_search(*args, duplicates=False):
	global_word_map = {}
	for arg in args:
		if(not arg):
			continue
		word_map = {}
		if(isinstance(arg, str)):
			arg = [arg]
		for words in map(lambda x: NON_ALPHA_GROUPS_REGEX.split(x.lower()), arg):
			for word in words:
				if(not word):
					continue
				key = word[:5]
				existing_words_of_key = word_map.get(key)
				if(not existing_words_of_key):
					word_map[key] = existing_words_of_key = []
				existing_words_of_key.append(word)
		for _key, words in word_map.items():
			_words = global_word_map.get(_key)
			# when existing matching words list
			# has more than current arg, ignore
			if(_words and len(_words) > len(words)):
				continue
			global_word_map[_key] = words

	ret = []
	for key, vals in global_word_map.items():
		ret.extend(vals)

	return ret if duplicates else list(set(ret))


# returns None when there are exceptions instead of throwing
def ignore_exceptions(*exceptions):
	if(
		exceptions
		and isinstance(exceptions[0], types.FunctionType)
	):
		# using as a simple decorator
		func = exceptions[0]

		def new_func(*args, **kwargs):
			try:
				func(*args, **kwargs)
			except Exception as ex:
				LOG_WARN("ignoring_exception", func=func.__name__, exception=str(ex))
			return None

		new_func._original = getattr(func, "_original", func)
		return new_func
	else:
		# decorator with exceptions arg
		exceptions = tuple(exceptions) if exceptions else (Exception,)

		def decorator(func):
			def new_func(*args, **kwargs):
				try:
					func(*args, **kwargs)
				except Exception as ex:
					if(not isinstance(ex, exceptions)):
						raise ex
					LOG_WARN("ignoring_exception", func=func.__name__, exception=str(ex))
				return None

			new_func._original = getattr(func, "_original", func)
			return new_func
		return decorator


# r etries all exceptions or specific given expections only
# backoff = 1 => exponential sleep
# max_time milliseconds for exception to sleep, not counts the func runtime
def retry(num_retries, ignore_exceptions=None, max_time=5000):
	num_retries = max(2, num_retries)
	sleep_time_on_fail = max_time / num_retries
	ignore_exceptions = tuple(ignore_exceptions) if ignore_exceptions else (Exception,)

	def decorator(func):
		def new_func(*args, **kwargs):
			retry_count = 0
			while(retry_count < num_retries):
				try:
					return func(*args, **kwargs)
				except Exception as ex:
					if(not isinstance(ex, ignore_exceptions)):
						raise ex
					LOG_WARN("retrying", func=func.__name__, exception=str(ex))
					sleep(sleep_time_on_fail / 1000)
				retry_count += 1
			return None

		new_func._original = getattr(func, "_original", func)
		return new_func
	return decorator


def original_function(func):
	_original = getattr(func, "_original", _OBJ_END_)
	if(_original is not _OBJ_END_):
		return _original

	while(True):
		func_wrapped = getattr(func, "__wrapped__", _OBJ_END_)
		if(func_wrapped is _OBJ_END_):
			break
		func = func_wrapped

	return func


def empty_func():
	pass


# 1. Find appropriate partitioned task queue
# 2. If not found, create a new one, and start a worker thread for this queue
# 3. The worker thread waits for tasks to be added to the queue and wakes up
class PartitionedTasksRunner:
	idle_pt_queues = None  # for reusing, push a task, and thread continues
	partitioned_task_queues = None  # running queues
	started = False

	def __init__(self, max_parallel=10000) -> None:
		super().__init__()
		self.idle_pt_queues = Queue()
		self.partitioned_task_queues = {}
		self.max_parallel = max_parallel
		self.pt_runners_to_join = []
		self.is_processing = True

	def worker(self, _queue):
		LOG_DEBUG("partitioned_tasks_runner", msg="tasks runner worker started")
		while(_task := _queue.get()):  # idenfinitely waits for task until None (flushed)
			partition_key, queued_at, log_trace_id, func, args, kwargs = _task
			log_ctx._trace_id = log_trace_id  # set the trace id
			# call the function
			now_millis = cur_ms()
			if((delay := now_millis - queued_at) > 4000):
				LOG_WARN("background_task_delay", delay=delay, func=func.__name__, partition_key=partition_key)

			try:
				func(*args, **kwargs)
			except Exception as ex:
				LOG_ERROR(
					"background_task", partition_key=partition_key,
					desc=str(ex), stracktrace_string=traceback.format_exc()
				)

			# signal the main thread to ask for more tasks
			if(_queue.empty()):
				self.partitioned_task_queues.pop(partition_key, None)
				self.idle_pt_queues.put(_queue)

		LOG_DEBUG("partitioned_tasks_runner", msg="tasks runner worker stopped")

	# will stop taking new tasks and wait for all existing tasks to finish
	def stop(self):
		self.is_processing = False
		for pt_runner, pt_queue in self.pt_runners_to_join:
			pt_queue.put(None)  # flush
			pt_runner.join()
		self.idle_pt_queues = Queue()  # the previous queues are no longer running
		self.pt_runners_to_join.clear()

	def submit_task(self, partition_key, func, args, kwargs, max_backlog=None, timeout=None):
		if(not self.is_processing):
			# CHECK IF THIS SPANED FROM A BACKGROUND TASK
			LOG_ERROR(
				"background_tasks_runner", msg="cannot process new tasks in background as it's stopped. running them in main thread",
				stack_trace_string=traceback.format_exc()
			)
			func(*args, **kwargs)
			return

		now_millis = cur_ms()
		if(partition_key is None):
			partition_key = str(now_millis % 100)  # 50 parallel partitions

		pt_queue = self.partitioned_task_queues.get(partition_key)  # partitioned task queue
		# reuse from idle_pt_queues if possible
		if(not pt_queue):
			try:
				if(timeout and len(self.pt_runners_to_join) >= self.max_parallel):
					pt_queue = self.idle_pt_queues.get(timeout=timeout)  # wait for some timeout if given
				else:
					# don't wait fail fast, so we can create a runner
					pt_queue = self.idle_pt_queues.get(block=False)
			except QueueEmptyException:
				if(len(self.pt_runners_to_join) >= self.max_parallel):
					raise Exception(f"Maximum parallel executions in progress : {self.max_parallel}")
				pt_queue = Queue()
				_thread = Thread(  # each queue has a corresponding thread
					target=self.worker,
					args=(pt_queue,)
				)
				_thread.start()
				self.pt_runners_to_join.append((_thread, pt_queue))
			# mark it as being used by this partion key
			self.partitioned_task_queues[partition_key] = pt_queue

		if(max_backlog is not None and len(pt_queue) >= max_backlog):
			raise Exception("maximum backlog in patition reached")

		# submit task to the partitioned queue
		pt_queue.put((partition_key, now_millis, log_ctx.trace_id, func, args, kwargs))


partitioned_tasks_runner = PartitionedTasksRunner()


# submit a task:func to a partition
# parition is used when you want them to execute in the
# same order as submitted
def submit_background_task(partition_key, func, *args, **kwargs):
	partitioned_tasks_runner.submit_task(partition_key, func, args, kwargs)


# decorator to be used for a short io tasks to be
# run in background
def background_task(func):
	def wrapper(*args, **kwargs):
		# spawn the thread
		partitioned_tasks_runner.submit_task(None, func, args, kwargs)
		return

	wrapper._original = getattr(func, "_original", func)
	return wrapper


def background_task_partitioned(partition_key):
	partition_key_func = partition_key

	def _simple_partition_key_func(args, kwargs):
		return str(partition_key)

	if(not callable(partition_key)):
		partition_key_func = _simple_partition_key_func

	def _wrapper(func):
		def wrapper(*args, **kwargs):
			# spawn the thread
			partitioned_tasks_runner.submit_task(
				partition_key_func(args, kwargs) if partition_key_func else None,  # parition key
				func, args, kwargs
			)
			return
		wrapper._original = getattr(func, "_original", func)
		return wrapper
	return _wrapper


@events.register_listener(["blaster_exit5"])
def exit_5():
	partitioned_tasks_runner.stop()  # stop all background runners


# calls a function after the function returns given by argument after
def call_after_func(func):

	if(isinstance(func, str)):
		# after_func => take from args named by func
		def decorator(func):
			def new_func(*args, **kwargs):
				after_func = kwargs.pop(func, None)
				ret = func(*args, **kwargs)
				after_func and after_func()
				return ret

			new_func._original = getattr(func, "_original", func)
			return new_func
		return decorator

	else:
		def new_func(*args, after=None, **kwargs):
			ret = func(*args, **kwargs)
			after and after()
			return ret

		new_func._original = getattr(func, "_original", func)
		return new_func


def __iter_subclasses(cls, seen, reverse) -> iter:
	if(cls in seen):
		return
	seen.add(cls)
	if(not reverse):
		yield cls
	for sub_class in cls.__subclasses__():
		yield from __iter_subclasses(sub_class, seen, reverse)
	if(reverse):
		yield cls


def all_subclasses(cls, reverse=False, ignore_initial=True) -> iter:
	_iter = __iter_subclasses(cls, set(), reverse)
	if(ignore_initial):
		next(_iter)
	return _iter


def read_rows_from_url(url, csv_delimiter=",") -> iter:
	if(url.startswith("http")):
		with requests.get(url, stream=True) as resp:
			if(url.endswith(".xlsx")):
				xls_file = BytesIO()
				for chunk in resp.iter_content(chunk_size=_16KB_):
					xls_file.write(chunk)  # unfortunately we read the whole file, may be we can cut off at 10MB or something
				xls_file.seek(0)
				excel_sheet = openpyxl.load_workbook(xls_file, read_only=True, data_only=True).active
				for row in excel_sheet.iter_rows(values_only=True):
					yield row
			elif(url.endswith(".xls")):
				sheet = xlrd.open_workbook(file_contents=resp.content).sheet_by_index(0)
				for cells in sheet.get_rows():
					yield [repr(c.value) for c in cells]
			else:
				for row in csv.reader(resp.iter_lines(decode_unicode=True), delimiter=csv_delimiter):
					yield row
	else:
		with open(url, "rb") as file:
			if(url.endswith(".xlsx")):
				excel_sheet = openpyxl.load_workbook(file, read_only=True, data_only=True).active
				for row in excel_sheet.iter_rows(values_only=True):
					yield row
			elif(url.endswith(".xls")):
				sheet = xlrd.open_workbook(file_contents=resp.content).sheet_by_index(0)
				for cells in sheet.get_rows():
					yield [repr(c.value) for c in cells]
			else:
				for row in csv.reader(TextIOWrapper(file), delimiter=csv_delimiter):
					yield row


def to_csv_bytes(rows):
	csv_file = StringIO()
	csv_writer = csv.writer(csv_file, quoting=csv.QUOTE_MINIMAL)
	csv_writer.writerows(rows)
	return csv_file.getvalue().encode("utf-8")


# print debugging info for networks request called with requests
def debug_requests_on():
	import logging
	'''Switches on logging of the requests module.'''
	HTTPConnection.debuglevel = 3
	logging.basicConfig()
	logging.getLogger().setLevel(logging.DEBUG)
	requests_log = logging.getLogger("requests.packages.urllib3")
	requests_log.setLevel(logging.DEBUG)
	requests_log.propagate = True


def debug_requests_off():
	'''Switches off logging of the requests module, might be some side-effects'''
	import logging
	HTTPConnection.debuglevel = 0

	root_logger = logging.getLogger()
	root_logger.setLevel(logging.WARNING)
	root_logger.handlers = []
	requests_log = logging.getLogger("requests.packages.urllib3")
	requests_log.setLevel(logging.WARNING)
	requests_log.propagate = False


@contextlib.contextmanager
def debug_requests():
	'''Use with 'with'!'''
	debug_requests_on()
	yield
	debug_requests_off()


def cached_request(
	url, ignore_cache_read=False, cache_folder="/tmp/",
	as_string_buffer=False, _json=None, data=None, headers=None
):
	cache_hash = url
	if(_json):
		cache_hash += json.dumps(_json)
	elif(data):
		cache_hash += json.dumps(data)
	cache_hash = hashlib.md5(cache_hash.encode("utf-8")).hexdigest()
	cache_file_path = cache_folder + cache_hash
	if(ignore_cache_read or not os.path.isfile(cache_file_path)):
		if(_json or data):
			resp = requests.post(url, json=_json, data=data, headers=headers, stream=True)
		else:
			resp = requests.get(url, headers=headers, stream=True)
		try:
			with open(cache_file_path, "wb") as cache_file:
				for chunk in resp.iter_content():
					cache_file.write(chunk)
		except Exception:
			os.remove(cache_file_path)

	bytes_buffer = BytesIO(open(cache_file_path, "rb").read())
	if(as_string_buffer):
		return StringIO(bytes_buffer.read().decode())
	else:
		return bytes_buffer


def memoized_method(func):
	cache = {}

	def wrapper(self, *args, **kwargs):
		# need weakref otherwise cache will never be garbage collected
		key = (weakref.ref(self), args, frozenset(kwargs.items()))
		if key in cache:
			return cache[key]
		result = func(self, *args, **kwargs)
		cache[key] = result
		return result
	setattr(wrapper, "_original", func)
	wrapper.clear_cache = cache.clear
	return wrapper


def NON_NULL(*args):
	for a in args:
		if a is not None:
			return a
	return None


def NON_NULL_DICT(kv):
	has_null_values = False
	for k, v in kv.items():
		if v is None:
			has_null_values = True
			break
	if has_null_values:
		return {k: v for k, v in kv.items() if v is not None}
	return kv


# RATE LIMITING
class RateLimitingException(Exception):
	pass


cache_by_rate_limiting_type = {}


def ASSERT_RATE_PER_MINUTE(key, count=60, _type=""):
	cache = cache_by_rate_limiting_type.get(_type)
	if(not cache):
		cache = cache_by_rate_limiting_type[_type] \
			= ExpiringCache(1000000, ttl=10 * 60 * 1000)

	# rate limit the key
	now_millis = cur_ms()
	if(
		not (rpm := cache.get(key))
		or rpm["last"] < now_millis - MILLIS_IN_MINUTE
	):
		rpm = {"c": 0, "last": now_millis}
		cache[key] = rpm

	rpm["c"] += 1
	if(rpm["c"] > count):
		raise RateLimitingException(f"Rate limit exceeded for {key}")


def set_requests_default_args(**_kwargs):
	import requests
	from requests.adapters import HTTPAdapter

	class _HTTPAdapter(HTTPAdapter):
		def send(self, request, **kwargs):
			for key, default in _kwargs.items():
				if(key not in kwargs):
					kwargs[key] = default
			return super().send(request, **kwargs)

	session = requests.Session()
	adapter = _HTTPAdapter()  # Set your default timeout in seconds
	session.mount("http://", adapter)
	session.mount("https://", adapter)


def xmltodict(xml_node, attributes=False):
	t = etree.fromstring(
		xml_node.encode("utf-8"),
		parser=etree.XMLParser(recover=True)
	) if isinstance(xml_node, str) else xml_node
	if(t is None):
		return None
	_attributes = t.attrib if attributes else None
	d = {t.tag: {} if attributes else None}
	children = list(t)
	if(children):
		dd = {}
		for dc in filter(None, map(
			lambda n: xmltodict(n, attributes=attributes),
			children
		)):
			for k, v in dc.items():
				if k in dd:
					if isinstance(dd[k], list):
						dd[k].append(v)
					else:
						dd[k] = [dd[k], v]
				else:
					dd[k] = v
		d = {t.tag: dd}

	if(_attributes):
		d[t.tag].update(('@' + k, v) for k, v in _attributes.items())

	if(t.text):
		text = t.text.strip()
		if(children or _attributes):
			d[t.tag]['#text'] = text
		else:
			d[t.tag] = text
	return d


def mask_strings(val, mask_char="*"):
	if(isinstance(val, str)):
		val = (val[:2] + mask_char * (len(val) - 2)) if val else val
	elif(isinstance(val, (list, tuple))):
		val = [mask_strings(v, mask_char) for v in val]
	elif(isinstance(val, dict)):
		val = {k: mask_strings(v, mask_char) for k, v in val.items()}
	return val


def create_operator_tree(expression, operators: list):
	# split the expression by the operators list and also '(' and ')' and build the tokens
	# convert infix to postfix and then build the tree
	# operators are already in sorted order of precedence
	tokens = []
	current_token = ""
	OPERAND = 1
	OPERATOR = 2
	BRACES = 3

	for char in expression:
		if(char.isspace()):
			if current_token:
				if(tokens and tokens[-1][0] == OPERAND):
					tokens[-1] = (OPERAND, tokens[-1][1] + " " + current_token)
				else:
					tokens.append((OPERAND, current_token))
				current_token = ""
			continue
		if(char in "()"):
			if current_token:
				tokens.append((OPERAND, current_token))
				current_token = ""
			tokens.append((BRACES, char))
			continue

		current_token += char
		if current_token in operators:
			tokens.append((OPERATOR, current_token))
			current_token = ""

	if(current_token):
		if(tokens and tokens[-1][0] == OPERAND):
			tokens[-1] = (OPERAND, tokens[-1][1] + " " + current_token)
		else:
			tokens.append((OPERAND, current_token))

	precedence = {op: i for i, op in enumerate(operators)}
	tokens = [token[1] for token in tokens]  # filter out braces

	def infix_to_postfix(tokens):
		output = []
		stack = []
		for token in tokens:
			if(token in precedence):
				while(stack and stack[-1] != '(' and precedence[stack[-1]] >= precedence[token]):
					output.append(stack.pop())
				stack.append(token)
			elif(token == '('):
				stack.append(token)
			elif(token == ')'):
				while(stack and stack[-1] != '('):
					output.append(stack.pop())
				if(stack and stack[-1] == '('):
					stack.pop()
			else:
				# it's an operand
				output.append(token)
		while(stack):
			output.append(stack.pop())
		return output

	def build_tree(postfix_tokens):
		stack = []
		for token in postfix_tokens:
			if(token in precedence):
				right = stack.pop()
				left = stack.pop()
				node = {'op': token, 'left': left, 'right': right}
				stack.append(node)
			else:
				stack.append(token)
		return stack[0] if stack else None
	postfix_tokens = infix_to_postfix(tokens)
	tree = build_tree(postfix_tokens)
	return tree
